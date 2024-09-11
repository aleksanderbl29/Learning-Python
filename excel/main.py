import yaml
from tabulate import tabulate
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Font, colors, fills, Alignment

## Import data from YAML file
with open("excel/info.yaml", "r") as f:
    data = yaml.load(f, Loader=yaml.SafeLoader)

## Print data in table format
print(tabulate(data, headers="keys", tablefmt="grid"))

## Load workbook
wb = load_workbook(filename = 'excel/kursus.xlsx')
ws = wb["Forside"]

## Define djof color
djof_color = colors.Color(rgb = 'ff00405f')

## Define cells
cells = [
    ws.cell(row=13, column=3),  # C13
    ws.cell(row=14, column=3),  # C14
    ws.cell(row=15, column=3),  # C15
    ws.cell(row=17, column=3),  # C17
    ws.cell(row=18, column=3),  # C18
    ws.cell(row=19, column=3),  # C19
    ws.cell(row=21, column=3),  # C21
    ws.cell(row=22, column=3),  # C22
    ws.cell(row=23, column=3)   # C23
]


name_cells = cells[0::3]
education_cells = cells[1::3]
email_cells = cells[2::3]

print(name_cells)
print(education_cells)
print(email_cells)

name_font = Font(name = 'Arial', size = 20, color = 'ffffffff')
education_font = Font(name = 'Arial', size = 16, color = 'ffffffff')
email_font = Font(name = 'Arial', size = 11, color = 'ffffffff')

fill = fills.PatternFill(fill_type = 'solid', fgColor = djof_color)

for cell in cells:
    cell.fill = fill
    row = cell.row
    ws.row_dimensions[row].height = 26

for name in name_cells:
    index = name_cells.index(name)
    name.font = name_font
    name.value = data[index]["name"]
    row = name.row
else:
    print("All names have been updated")

for education in education_cells:
    index = education_cells.index(education)
    education.font = education_font
    education.value = data[index]["education"]
else:
    print("All educations have been updated")

for email in email_cells:
    index = email_cells.index(email)
    email.font = email_font
    email.value = data[index]["email"]
    email.hyperlink = data[index]["email"]
else:
    print("All emails have been updated")

wb.save("excel/kursus_output.xlsx")

